import xlrd
from openpyxl import load_workbook
import openpyxl

wb = openpyxl.Workbook()
wb = load_workbook(filename='/home/rkverma11254/Downloads/Kai/ops_dataset_for_testing.xlsx')
wb.create_sheet(index=1, title='City')
wb.create_sheet(index=2, title='State')
wb.create_sheet(index=3, title='Country')
wb.create_sheet(index=4, title='PostalCode')
wb.create_sheet(index=5, title='Address')
wb.create_sheet(index=6, title='URL')
wb.create_sheet(index=7, title='CompanyName')
State = wb.active
City_Excel = wb.get_sheet_by_name('City')
State_Excel = wb.get_sheet_by_name('State')
Country_Excel = wb.get_sheet_by_name('Country')
PostalCode_Excel = wb.get_sheet_by_name('PostalCode')
Address_Excel = wb.get_sheet_by_name('Address')
URL_Excel = wb.get_sheet_by_name('URL')
Company_Name_Excel = wb.get_sheet_by_name('CompanyName')


workbook1 = xlrd.open_workbook('/home/rkverma11254/Downloads/Kai/ops_dataset_for_testing.xlsx')
workbook2 = xlrd.open_workbook('/home/rkverma11254/Downloads/Kai/preetam_dataset_for_testing.xls')
workshee1 = workbook1.sheet_by_name('1')
workshee2 = workbook2.sheet_by_name('1')

#====================================== City ===================================================================================================
for row in range(workshee1.nrows):
    data1 = str(workshee1.cell_value(row,9))
    data2 = str(workshee2.cell_value(row,42))
    if data1 == data2:
        pass
    else:
        with open("/home/rkverma11254/Music/City.txt", 'a+') as City:
            City.writelines("Row Number: {0} 'City': {1} 'N_City': {2} Status: {3} ".format(row, data1, data2, "Fail") + "\n")
        City.close()
# ================================ State========================================================================================================
for row in range(workshee1.nrows):
    data3 = str(workshee1.cell_value(row,10))
    data4 = str(str(workshee2.cell_value(row, 39)))
    if data3 == data4:
        pass
    else:
        with open("/home/rkverma11254/Music/State.txt", 'a+') as State:
            State.writelines("Row Number: {0} 'State': {1} 'N_State': {2}   Status: {3} ".format(row, data3, data4, "Fail") + "\n")
        State.close()
# =========================== PostalCode ======================================================================================================
for row in range(workshee1.nrows):
    data5 = str(workshee1.cell_value(row, 11))
    data6 = str(workshee2.cell_value(row, 38))
    if data5 == data6:
        pass
    else:
        with open("/home/rkverma11254/Music/PostalCode.txt", 'a+') as PostalCode:
            PostalCode.writelines("Row Number: {0} 'postal_code': {1} 'N_Postal_Code': {2}   Status: {3} ".format(row, data5, data6, "Fail") + "\n")
        PostalCode.close()
# ======================================== Country =============================================================================================
for row in range(workshee1.nrows):
    data7 = str(workshee1.cell_value(row, 13))
    data8 = str(workshee2.cell_value(row, 36))
    if data7 == data8:
        pass
    else:
        with open("/home/rkverma11254/Music/Country.txt", 'a+') as Country:
            Country.writelines("Row Number: {0} 'Country': {1} 'N_Country': {2}   Status: {3} ".format(row, data7, data8, "Fail") + "\n")
    Country.close()
#===================================== Address =================================================================================================
for row in range(workshee1.nrows):
    data9 = str(workshee1.cell_value(row, 6))
    data10 = str(workshee2.cell_value(row, 40))
    if data9 == data10:
        pass
    else:
          with open("/home/rkverma11254/Music/Address.txt", 'a+') as Address:
              Address.writelines("Row Number: {0} 'Address1': {1} 'N_Street_Address_1': {2}   Status: {3} ".format(row, data9, data10, "Fail") + "\n")
          Address.close()
#================================Company Name ==========================================================================================
for row in range(workshee1.nrows):
    data11 = str(workshee1.cell_value(row, 5))
    data12 = str(workshee2.cell_value(row, 35))
    if data11 == data12:
        pass
    else:
        with open("/home/rkverma11254/Music/Company.txt", 'a+') as Company:
            Company.writelines("Row Number: {0} 'Updated_Party_Name': {1} 'N_Company_Name': {2}   Status: {3} ".format(row, data11, data12, "Fail") + "\n")
        Company.close()
# =========================Domain==========================================================================================================
for row in range(workshee1.nrows):
    data13 = str(workshee1.cell_value(row, 14))
    data14 = str(workshee2.cell_value(row, 34))
    if data13 == data14:
        pass
    else:
        with open("/home/rkverma11254/Music/Domain.txt", 'a+') as Domain:
            Domain.writelines("Row Number: {0} 'URL': {1} 'N_Company_Name': {2}   Status: {3} ".format(row, data13, data14, "Fail") + "\n")
        Domain.close()

wb.close()


